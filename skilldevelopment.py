import os
import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from google.cloud import vision

# Set the path to the Google Cloud Vision API key
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r'C:\Users\HP\Downloads\sharp-airway-436916-t0-.json'

# Initialize Google Cloud Vision Client
def initialize_vision_client():
    return vision.ImageAnnotatorClient()

# Extract text using Google Cloud Vision
def extract_text_from_image(image_path, client):
    with open(image_path, 'rb') as image_file:
        content = image_file.read()
    
    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations

    if response.error.message:
        raise Exception(f'Error from Google Cloud Vision: {response.error.message}')
    
    extracted_text = texts[0].description if texts else ""
    print(f"\nExtracted text from image '{image_path}':\n{extracted_text}")
    
    return extracted_text

# Extract 'Grant Head' based on "TEQIP"
def extract_grant_head(text):
    # Convert text to uppercase for case-insensitive comparison
    upper_text = text.upper()

    # Check for the presence of 'TEQIP', 'TEA', or 'TECIP'
    if 'TEQIP' in upper_text or 'TEA' in upper_text or 'TE QIP' in upper_text or'TERIP' in upper_text or 'TEP' in upper_text:
        return 'TEQIP'
    
    return 'Not mentioned'

# Extract date using regex, accepting dd/mm and dd/mm/yyyy formats
def extract_date(text):
    date_pattern = r'\b(?:[0-3]?\d[-/][01]?\d(?:[-/]\d{2,4})?)\b'
    match = re.search(date_pattern, text)
    if match:
        return match.group(0)
    return None

# Extract 'Payment Made To' and handle "party payment TO SHRI MIS"




def extract_payment_made_to(text):
    # Try to match text after "payment to be made", "AV/s payment", "AU's payment", or "party payment TO SHRI"
    match_1 = re.search(r'payment (?:to(?:mado)?|to be made)\s*(?:A/s|AV/s|AU\'s|M/S|MIS|A\'s|Mr|Mrs|Ms|Dr)?[., ]*\s*([A-Za-z\s.,]+)', text, re.IGNORECASE)
    match_2 = re.search(r'party payment TO SHRI\s*(?:A/s|AV/s|AU\'s|M/S|MIS|A\'s|Mr|Mrs|Ms|Dr)?[., ]*\s*([A-Za-z\s.,]+)', text, re.IGNORECASE)
    match_3 = re.search(r'(?:A/s|AV/s|AU\'s)\s*(?:Mr|Mrs|Ms|Dr)?[., ]*\s*([A-Za-z\s.,]+)', text, re.IGNORECASE)

    # Extract name from the matches
    name_1 = match_1.group(1).strip() if match_1 else None
    name_2 = match_2.group(1).strip() if match_2 else None
    name_3 = match_3.group(1).strip() if match_3 else None

    # Choose the longest name to avoid partial matches
    name = max((n for n in [name_1, name_2, name_3] if n), key=len, default=None)

    if name:
        # Remove unnecessary words like "A/s", "AU's", "M/S", "Mr.", etc.
        name = re.sub(r'\b(?:A/s|AV/s|AU\'s|M/S|MIS|A\'s|Mr\.|Mrs\.|Ms\.|Dr\.)\b', '', name, flags=re.IGNORECASE).strip()

        # Remove any trailing unintended keywords like "Nos. Bs" or "End Nos. Bills"
        name = re.sub(r'(End Nos\. Bills|Nos\. Bs)', '', name, flags=re.IGNORECASE).strip()

        # Stop at any backslash, newline, or similar characters
        name = name.split('\\')[0].split('\n')[0].strip()

        return name

    return None


# Extract 'Total Amount'
def extract_total_amount(text):
    # Find all matches for 'Grand Total Rs.' or 'Total Rs.' followed by numbers (including commas),
    # allowing for optional whitespace, newlines, or bullet points
    matches = re.findall(r'(Grand Total Rs\.|Total Rs\.)\s*[\n]*\s*[\u2022]*\s*([0-9,]+(?:\.[0-9]+)?)', text, re.IGNORECASE)

    if matches:
        # Extract the last match (in case there are multiple amounts, we want the final one, usually Grand Total)
        last_match = matches[-1][1]  # Get the number part from the last match
        
        # Remove any commas from the amount string and convert to float to handle decimals
        amount = float(last_match.replace(',', ''))  
        return int(amount)  # Return as integer (10000 instead of 10000.00)

    return None

# Extract purpose based on specific keywords
def extract_purpose(text):
    text = text.lower()
    if 'industry' in text or 'indies' in text:
        return 'One day Industry Academia Meet'
    if 'industrial' in text:
        return 'Industrial Visit'
    if 'travel' in text:
        return 'Travelling expense'
    if 'tuition' in text:
        return 'PHD tuition fee reimbursement'
    if 'phit' in text:
        return 'Photocopy fee for semester'
    if 'fdp' in text:
        return 'Online Course IIT Roorkee'
   
    if 'netel' in text and 'bayramming' in text:
        return 'NPTEL course for Programming, Data Structure and Algorithm'
    if 'nptel' in text and 'tup' in text:
        return 'NPTEL course for Deep Learning'
    if 'nptel' in text and 'networks' in text:
        return 'NPTEL course for Computer Network and Internet Protocol'
    if 'nptel' in text and 'exam' in text:
        return 'NPTEL exam fee reimbursement'
    if 'nptel' in text and 'python' in text:
        return 'Python for Data science online course'
    if 'nptel' in text:
        return 'NPTEL course on Introduction to Machine Learning'
    return None

# Extract 'College Name' based on presence of 'SHRI' in text
def extract_college_name(text):
    match = re.search(r'^(.*SHRI.*)$', text, re.IGNORECASE | re.MULTILINE)
    if match:
        return match.group(1).strip()
    return None

# Parse extracted text to get bill data
def parse_bill_data(text):
    bill_data = {
        "Payment Made To": extract_payment_made_to(text),
        "Grant Head": extract_grant_head(text),
        "Date": extract_date(text),
        "Total Amount": extract_total_amount(text),
        "Purpose": extract_purpose(text),
        "College Name": extract_college_name(text)
    }

    print("Parsed Bill Data:", bill_data)
    return bill_data

# Organize parsed data into a DataFrame
def organize_data(bill_data):
    df = pd.DataFrame([bill_data])
    print("\nOrganized DataFrame:\n", df)
    return df

# Export DataFrame to Excel, appending new data without erasing old data
def export_to_excel(df, excel_path, create_new=False):
    if 'Payment Made To' in df.columns:
        df['Payment Made To'] = df['Payment Made To'].str.replace(r'\b(?:Mr\.|Mrs\.)\b', '', regex=True).str.strip()

    # If create_new is True, we don't try to load an existing Excel file
    if create_new:
        final_df = df
    else:
        try:
            existing_df = pd.read_excel(excel_path)
            final_df = pd.concat([existing_df, df], ignore_index=True)
        except FileNotFoundError:
            final_df = df

    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
        final_df.to_excel(writer, index=False)

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    for column in sheet.columns:
        sheet.column_dimensions[get_column_letter(column[0].column)].width = 30
    
    workbook.save(excel_path)
    print(f"Data exported to {excel_path}")


# Main function to process bill images
def main(image_paths, excel_path, create_new=False):
    client = initialize_vision_client()
    all_data = []

    for image_path in image_paths:
        text = extract_text_from_image(image_path, client)
        bill_data = parse_bill_data(text)
        df = organize_data(bill_data)
        all_data.append(df)

    if all_data:
        final_df = pd.concat(all_data, ignore_index=True)
        export_to_excel(final_df, excel_path, create_new=create_new)
    else:
        print("No data to export.")

# Example usage
if __name__ == "__main__":
    image_paths = [r'C:\Users\HP\Desktop\bills\Screenshot 2024-10-06 211947.png']
    excel_path = r'C:\Users\HP\Desktop\bill2.xlsx'
    create_new_excel = False  # Set to True if you want to create a new Excel file
    main(image_paths, excel_path, create_new=create_new_excel)
